import openpyxl
import sys

def copy_first_sheet(source_file, destination_file):
    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_file)
    # Get the first sheet from the source workbook
    first_sheet_name = source_workbook.sheetnames[0]
    source_sheet = source_workbook[first_sheet_name]

    # Load the destination workbook
    destination_workbook = openpyxl.load_workbook(destination_file)

    # Copy the sheet to the destination workbook
    target_sheet = destination_workbook.copy_worksheet(source_sheet)

    # Save the destination workbook
    destination_workbook.save(destination_file)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python script.py source_file destination_file")
    else:
        source_file = sys.argv[1]
        destination_file = sys.argv[2]
        copy_first_sheet(source_file, destination_file)
